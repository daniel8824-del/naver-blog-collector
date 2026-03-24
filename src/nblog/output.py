"""블로그 수집 결과 출력 포맷터 - 터미널, CSV, JSON, Markdown, Excel"""

import csv
import json
import io
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table
from rich.panel import Panel

console = Console()


def _prepare_output_path(filepath: str) -> Path:
    """저장 경로의 부모 디렉토리를 미리 생성."""
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def print_results(articles: list, query: str = ""):
    """Rich 테이블로 터미널에 블로그 수집 결과 출력."""
    if not articles:
        console.print("[yellow]검색 결과가 없습니다.[/yellow]")
        return

    console.print()
    console.print(Panel(
        f"[bold]검색어:[/bold] {query}\n"
        f"[bold]결과:[/bold] {len(articles)}건  "
        f"[dim]{datetime.now().strftime('%Y-%m-%d %H:%M')}[/dim]",
        title="[bold blue]블로그 수집 결과[/bold blue]",
        border_style="blue",
    ))

    table = Table(show_header=True, header_style="bold cyan", show_lines=True)
    table.add_column("#", style="dim", width=4, justify="right")
    table.add_column("제목", width=40)
    table.add_column("블로거", width=15)
    table.add_column("날짜", width=12)
    table.add_column("본문길이", width=10, justify="right")
    table.add_column("추출방법", width=10)

    for i, a in enumerate(articles, 1):
        title = a.get("title", "(제목 없음)")
        blogger = a.get("bloggerName", "")
        postdate = a.get("postdate", "")
        content = a.get("content", "")
        length = a.get("content_length", len(content))
        method = a.get("method", "")

        table.add_row(
            str(i),
            title,
            blogger,
            postdate,
            f"{length:,}자",
            method,
        )

    console.print(table)
    console.print()


def to_csv(articles: list, filepath: str, query: str = ""):
    """UTF-8 BOM CSV 파일로 저장."""
    path = _prepare_output_path(filepath)

    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["키워드", "타이틀", "본문", "블로거", "링크", "날짜", "URL"])

        for a in articles:
            writer.writerow([
                a.get("keyword", query),
                a.get("title", ""),
                a.get("content", ""),
                a.get("bloggerName", ""),
                a.get("bloggerLink", ""),
                a.get("postdate", ""),
                a.get("url", ""),
            ])

    console.print(f"[green]CSV 저장: {filepath}[/green]")


def to_txt(articles: list, filepath: str, query: str = ""):
    """전체 정보 텍스트 파일로 저장. 테이블 형태 + 본문."""
    path = _prepare_output_path(filepath)
    lines = []
    lines.append(f"네이버 블로그 수집 결과: {query}")
    lines.append(f"수집일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"총 {len(articles)}건")
    lines.append("=" * 70)

    for i, a in enumerate(articles, 1):
        title = a.get("title", "")
        blogger = a.get("bloggerName", "")
        postdate = a.get("postdate", "")
        url = a.get("url", "")
        content = a.get("content", "")
        length = a.get("content_length", len(content))

        lines.append(f"\n[{i}] {title}")
        lines.append(f"    블로거: {blogger}")
        lines.append(f"    날짜:   {postdate}")
        lines.append(f"    URL:    {url}")
        lines.append(f"    글자수: {length:,}자")
        lines.append("-" * 70)
        lines.append(content)
        lines.append("=" * 70)

    text = "\n".join(lines) + "\n"
    path.write_text(text, encoding="utf-8")
    console.print(f"[green]TXT 저장: {filepath}[/green]")


def to_excel(articles: list, filepath: str, query: str = ""):
    """openpyxl로 Excel 파일 저장. 시트명=키워드, 컬럼 너비 자동조절."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = query[:31] if query else "블로그 수집"

    # 스타일
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["키워드", "타이틀", "본문", "블로거", "링크", "날짜", "URL"]
    col_widths = [12, 40, 80, 15, 30, 12, 50]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    # 데이터
    body_font = Font(name="맑은 고딕", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for i, a in enumerate(articles, 1):
        row = i + 1
        values = [
            a.get("keyword", query),
            a.get("title", ""),
            a.get("content", "")[:32000],
            a.get("bloggerName", ""),
            a.get("bloggerLink", ""),
            a.get("postdate", ""),
            a.get("url", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = thin_border

    # 필터 + 틀 고정
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(articles) + 1}"
    ws.freeze_panes = "A2"

    path = _prepare_output_path(filepath)
    wb.save(str(path))
    console.print(f"[green]Excel 저장: {filepath}[/green]")


def to_json(articles: list, filepath: str):
    """JSON 파일로 저장."""
    data = {
        "collected_at": datetime.now().isoformat(),
        "count": len(articles),
        "articles": articles,
    }
    json_text = json.dumps(data, ensure_ascii=False, indent=2)

    path = _prepare_output_path(filepath)
    path.write_text(json_text, encoding="utf-8")
    console.print(f"[green]JSON 저장: {filepath}[/green]")


def to_markdown(articles: list, filepath: str):
    """마크다운 테이블 파일로 저장."""
    lines = [
        "| # | 제목 | 블로거 | 날짜 | 본문길이 | 추출방법 |",
        "|---|------|--------|------|----------|----------|",
    ]

    for i, a in enumerate(articles, 1):
        title = a.get("title", "").replace("|", "\\|")
        blogger = a.get("bloggerName", "").replace("|", "\\|")
        postdate = a.get("postdate", "")
        length = a.get("content_length", 0)
        method = a.get("method", "")

        lines.append(f"| {i} | {title} | {blogger} | {postdate} | {length:,} | {method} |")

    md_text = "\n".join(lines) + "\n"

    path = _prepare_output_path(filepath)
    path.write_text(md_text, encoding="utf-8")
    console.print(f"[green]Markdown 저장: {filepath}[/green]")
